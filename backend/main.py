from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import openpyxl
import io

from backend.nsr10.derivas import procesar_derivas, limite_deriva_cm
from backend.nsr10.torsion import evaluar_torsion_proyecto

app = FastAPI(
    title="StructurApp API",
    description="Sistema de verificación estructural NSR-10",
    version="0.1.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def raiz():
    return {"app": "StructurApp", "version": "0.1.0", "estado": "activo"}

@app.post("/api/v1/derivas/calcular")
async def calcular_derivas(archivo: UploadFile = File(...)):
    if not archivo.filename.endswith(".xlsx"):
        raise HTTPException(400, "El archivo debe ser .xlsx")

    contenido = await archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb["Joint Drifts"] if "Joint Drifts" in wb.sheetnames else wb.active

    NODOS_EXTREMOS = {"1", "9", "25", "85"}
    CASOS_DERIVA = {"Espectro X (Deriva)", "Espectro Y (Deriva)"}

    registros_raw = []
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if fila[0] is None:
            continue
        story = fila[0]
        label = str(fila[1]) if fila[1] is not None else ""
        output_case = fila[3]
        drift_x = fila[9]
        drift_y = fila[10]
        if label not in NODOS_EXTREMOS:
            continue
        if output_case not in CASOS_DERIVA:
            continue
        if drift_x is None or drift_y is None:
            continue
        registros_raw.append({
            "piso": story,
            "nodo": label,
            "caso_carga": output_case,
            "drift_x": float(drift_x),
            "drift_y": float(drift_y),
        })

    if not registros_raw:
        raise HTTPException(422, "No se encontraron datos en el archivo")

    derivas = procesar_derivas(registros_raw)
    pisos_torsion = _construir_pares_torsion(derivas)
    torsion = evaluar_torsion_proyecto(pisos_torsion)

    return {
        "resumen": {
            "pisos_analizados": len({r["piso"] for r in derivas}),
            "nodos_analizados": len({r["nodo"] for r in derivas}),
            "limite_deriva_cm": limite_deriva_cm(),
            "phi_p_global": torsion["phi_p_global"],
            "clasificacion": torsion["clasificacion"],
            "tiene_1aP": torsion["tiene_1aP"],
            "tiene_1bP": torsion["tiene_1bP"],
        },
        "derivas": derivas,
        "torsion": torsion["detalle_por_piso"],
    }

def _construir_pares_torsion(derivas):
    from collections import defaultdict
    agrupado = defaultdict(dict)
    for r in derivas:
        key = (r["piso"], r["caso_carga"])
        agrupado[key][r["nodo"]] = r["deriva_combinada_cm"]

    pares_config = {
        "Espectro X (Deriva)": [("25", "1"), ("85", "9")],
        "Espectro Y (Deriva)": [("25", "85"), ("1", "9")],
    }

    resultado = []
    for (piso, caso), nodos in agrupado.items():
        dir_sismo = "X" if "X" in caso else "Y"
        for nodo_a, nodo_b in pares_config.get(caso, []):
            if nodo_a in nodos and nodo_b in nodos:
                resultado.append({
                    "piso": piso,
                    "direccion_sismo": dir_sismo,
                    "par_nodos": f"{nodo_a} vs {nodo_b}",
                    "delta1_cm": nodos[nodo_a],
                    "delta2_cm": nodos[nodo_b],
                })
    return resultado

app.mount("/frontend", StaticFiles(directory="frontend"), name="frontend")